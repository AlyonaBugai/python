import telebot
import openpyxl

access_token = "952661047:AAGRvG7gqnV-EfvN7n7291baHtIbiuJMQQw"
bot = telebot.TeleBot(access_token)

wb = openpyxl.load_workbook(filename='rozklad.xlsx')


# apr tv dif alg


def get_mon(sheet):
    ws = wb[sheet]
    column_mon = ws['A']
    vals = []
    for i in range(len(column_mon)):
        vals.append(column_mon[i].value)
    return '\n'.join(map(str, vals))


def get_tue(sheet):
    ws = wb[sheet]
    column_tue = ws['B']
    vals = []
    for i in range(len(column_tue)):
        vals.append(column_tue[i].value)
    return '\n'.join(map(str, vals))


def get_wed(sheet):
    ws = wb[sheet]
    column_wed = ws['C']
    vals = []
    for i in range(len(column_wed)):
        vals.append(column_wed[i].value)
    return '\n'.join(map(str, vals))


def get_thu(sheet):
    ws = wb[sheet]
    column_thu = ws['D']
    vals = []
    for i in range(len(column_thu)):
        vals.append(column_thu[i].value)
    return '\n'.join(map(str, vals))


def get_fri(sheet):
    ws = wb[sheet]
    column_fri = ws['E']
    vals = []
    for i in range(len(column_fri)):
        vals.append(column_fri[i].value)
    return '\n'.join(map(str, vals))


group = ''
day = ''
schedule = ''

keyboard1 = telebot.types.ReplyKeyboardMarkup(True, True)
keyboard1.row('1', '2', '3', '4')
keyboard2 = telebot.types.ReplyKeyboardMarkup(True, True)
keyboard2.row('mon', 'tue', 'wed', 'thu', 'fri')


@bot.message_handler(content_types=['text'])
def start(message):

    if message.text == '/start':
        bot.send_message(message.from_user.id, 'Оберіть групу:'
                                               '\n Апроксимація і стохастика: 1'
                                               '\n Ймовірність, інформація, обробка даних: 2'
                                               '\n Диференціальні рівняння: 3'
                                               '\n Математичні методи обчислень, захисту інформації: 4\n', reply_markup=keyboard1)
        bot.register_next_step_handler(message, get_group)
    else:
        bot.send_message(message.from_user.id, 'Напиши /start')


def get_group(message):
    global group
    group = message.text
    sheet1 = 1
    sheet2 = 2
    sheet3 = 3
    sheet4 = 4

    if group == "1":
        group = sheet1
    elif group == "2":
        group = sheet2
    elif group == "3":
        group = sheet3
    elif group == "4":
        group = sheet4

    group = message.text
    bot.send_message(message.from_user.id, 'Оберіть день тижня:', reply_markup=keyboard2)
    bot.register_next_step_handler(message, day_w)


def day_w(message):
    global day
    day = message.text
    global schedule
    if day == "mon":
        schedule = get_mon(group)

    elif day == "tue":
        schedule = get_tue(group)

    elif day == "wed":
        schedule = get_wed(group)

    elif day == "thu":
        schedule = get_thu(group)

    elif day == "fri":
        schedule = get_fri(group)
    bot.send_message(message.from_user.id, text=schedule)
bot.polling(none_stop=True)


